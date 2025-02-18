import os
import tempfile
import zipfile
import argparse
from openpyxl import Workbook, load_workbook
from datetime import datetime


def read_leave_data(leave_file_path):
    """Reads the leave data from an Excel file and returns it as a dictionary."""
    leave_data = {}
    try:
        wb = load_workbook(leave_file_path)
        sheet = wb.active  # Access the active sheet in the workbook
        for row in sheet.iter_rows(min_row=2, values_only=True):
            leave_date = row[0]  # Read the date from the first column
            if isinstance(leave_date, str):  # If the date is in string format
                leave_date = datetime.strptime(leave_date, '%Y-%m-%d')  # Convert it to datetime object
            elif isinstance(leave_date, datetime):  # If it's already a datetime object
                leave_date = leave_date  # No conversion needed
            else:
                continue  # Skip rows with invalid dates

            emails = row[1] if row[1] else ""  # Read the emails from the second column, default to empty if None
            email_set = {email.strip() for email in emails.split(",") if email.strip()}  # Clean and convert to a set

            # Ensure the data is in the desired format
            if leave_date not in leave_data:
                leave_data[leave_date] = {'attendees': set()}  # Initialize the 'attendees' set
            leave_data[leave_date]['attendees'].update(email_set)  # Add emails to the 'attendees' set
    except Exception as e:
        print(f"Error reading leave data: {e}")
    return leave_data

def extract_zip(zip_file_path, extract_to):
    """Extracts the contents of a zip file to a specified directory."""
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        zip_ref.extractall(extract_to)


def generate_output_filename(meeting_name):
    """Generates a unique output filename by appending a number if the file already exists."""
    base_name = f'RnD_{meeting_name.replace(" ", "_")}_Data'
    output_file_name = f'{base_name}.xlsx'
    count = 1
    while os.path.exists(output_file_name):
        output_file_name = f'{base_name}_{count}.xlsx'
        count += 1
    return output_file_name


def compile_attendee_data(zip_file_path, meeting_name, start_date, end_date, email_list, leave_data):
    """Compiles meeting attendee data from Excel files"""
    result_dict = {}
    with tempfile.TemporaryDirectory() as temp_dir:
        extract_zip(zip_file_path, temp_dir)
        meeting_reports_dir = os.path.join(temp_dir, os.listdir(temp_dir)[0])
        for folder_name in os.listdir(meeting_reports_dir):
            folder_path = os.path.join(meeting_reports_dir, folder_name)
            if not os.path.isdir(folder_path):
                print("Not a valid directory.")
                continue
            try:
                folder_date = datetime.strptime(
                    folder_name.split(' ')[0], "%Y-%m-%d")
            except ValueError:
                print("Invalid folder name")
                continue
            if not (start_date <= folder_date <= end_date):
                continue
            if meeting_name not in folder_name:
                print(f"{meeting_name} not found in {folder_name}")
                continue  # Skip this folder if the meeting name is not found

            process_excel_files(folder_path, folder_date,
                                email_list, result_dict)

        # Merge result_dict with leave_data
        for date, leave_info in leave_data.items():
            if date in result_dict:
                # Filter attendees from leave_info to include only those in email_list
                filtered_attendees = leave_info['attendees'].intersection(set(email_list))
                # Merge the filtered attendees into the result_dict
                result_dict[date]['attendees'] = result_dict[date]['attendees'].union(filtered_attendees)

        # Pass the merged dictionary to subsequent functions
        save_to_excel(generate_output_filename(meeting_name).replace(
            '.csv', '.xlsx'), result_dict, email_list)
        individual_attendance_filename = generate_output_filename(
            f"{meeting_name}_Individual_Attendance")
        save_individual_attendee_percentages(individual_attendance_filename.replace(
            '.csv', '.xlsx'), result_dict, leave_data, email_list)


def process_excel_files(folder_path, folder_date, email_list, result_dict):
    """Process each Excel file based on column names."""
    for file_name in os.listdir(folder_path):
        if not file_name.endswith('.xlsx'):
            print("No xlsx files found in directory")
            continue
        file_path = os.path.join(folder_path, file_name)
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            # Get the header row to identify column indices
            header = [cell.value for cell in sheet[1]]
            email_index = header.index('Email') if 'Email' in header else None
            # Check if required columns are present
            if email_index is None:
                print(f"Missing required 'Email' column in file: {file_name}")
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                email = row[email_index]
                # Check if the email is in the provided email list
                if email in email_list:
                    result_dict.setdefault(folder_date, {'attendees': set()})['attendees'].add(email)
        except Exception as e:
            print(f"Error processing {file_path}: {e}")


def save_to_excel(output_file_name, result_dict, email_list):
    """Save the compiled results to an Excel file."""
    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.append(['Date', 'Attendee Emails', 'Percentage'])
    total_attendees_count = 0
    total_attendees = len(email_list)
    for date, data in result_dict.items():
        attendees_list = list(data['attendees'])
        attendees_count = len(attendees_list)
        attendance_percentage = (attendees_count / total_attendees * 100) if total_attendees > 0 else 0
        output_sheet.append([date.strftime('%Y-%m-%d'), ", ".join(attendees_list), f"{attendance_percentage:.2f}%"])
        total_attendees_count += attendees_count
    overall_percentage = (total_attendees_count / (len(result_dict) * total_attendees) * 100) if total_attendees > 0 else 0
    output_sheet.append([])
    output_sheet.append(['Total Percentage', '', f"{overall_percentage:.2f}%"])
    output_workbook.save(output_file_name)
    print(f"Compiled data saved to '{output_file_name}'")


def save_individual_attendee_percentages(output_file_name, result_dict, leave_data, email_list):
    """Save attendees and their attendance percentages to an Excel file."""
    
    print("Starting to calculate attendance percentages...")

    # Initialize the workbook and sheet
    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.append(['Attendee', 'Percentage'])

    # Dictionary to hold counts for each attendee
    attendee_counts = {}

    # Count attendance for each attendee across all meetings
    print("Counting attendance for each attendee...")
    for date, data in result_dict.items():
        print(f"Processing data for date: {date}")
        for attendee in data['attendees']:
            print(f"Adding attendee: {attendee}")
            attendee_counts[attendee] = attendee_counts.get(attendee, 0) + 1

    total_meetings = len(result_dict)
    print(f"Total number of meetings: {total_meetings}")

    # Calculate attendance percentage and save to Excel
    print("Calculating attendance percentages and saving to the file...")
    for attendee, count in attendee_counts.items():
        attendance_percentage = (count / total_meetings * 100) if total_meetings > 0 else 0
        print(f"{attendee} attended {count} meetings, which is {attendance_percentage:.2f}%")
        output_sheet.append([attendee, f"{attendance_percentage:.2f}%"])

    # Save the workbook
    output_workbook.save(output_file_name)
    print(f"Attendee percentages saved to '{output_file_name}'")


def main():
    """Main method"""
    parser = argparse.ArgumentParser(
        description='Compile attendee data from a ZIP file.')
    parser.add_argument('zip_file_path', type=str,
                        help='Path to the ZIP file containing meeting reports.')
    parser.add_argument('meeting_name', type=str, help='Name of the meeting.')
    parser.add_argument('start_date', type=str,
                        help='Start date for the reports in YYYY-MM-DD format.')
    parser.add_argument('end_date', type=str,
                        help='End date for the reports in YYYY-MM-DD format.')
    parser.add_argument('email_list', type=str,
                        help='Space-separated list of email addresses of attendees.')
    parser.add_argument('leave_file_path', type=str,
                        help='Path to the Excel file containing leave data.')
    args = parser.parse_args()
    zip_file_path = args.zip_file_path
    meeting_name = args.meeting_name
    start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
    end_date = datetime.strptime(args.end_date, '%Y-%m-%d')
    email_list = args.email_list.split()
    leave_data = read_leave_data(args.leave_file_path)
    print("This is leave data\n", leave_data)
    compile_attendee_data(zip_file_path, meeting_name,
                          start_date, end_date, email_list, leave_data)


if __name__ == '__main__':
    main()
